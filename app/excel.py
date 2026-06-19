"""Excel report builders (openpyxl) — port the useful formatting from the
Access VBA Excel-COM macros, minus the desktop-only chart automation."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

CURRENCY_FMT = "#,##0.00 $"
PCT_FMT = "0.00%"


def _autosize(ws, max_width: int = 60) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(w + 2, max_width)


def _write_sheet(ws, headers, rows, *, currency_cols=(), pct_cols=()):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(h) if isinstance(r, dict) else r[i] for i, h in enumerate(headers)])

    for name in currency_cols:
        if name in headers:
            idx = headers.index(name) + 1
            letter = get_column_letter(idx)
            for cell in ws[letter][1:]:
                cell.number_format = CURRENCY_FMT
    for name in pct_cols:
        if name in headers:
            idx = headers.index(name) + 1
            letter = get_column_letter(idx)
            for cell in ws[letter][1:]:
                cell.number_format = PCT_FMT

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _autosize(ws)


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- report builders --------------------------------------------------------

COST_HEADERS = [
    "Material", "Materialkurztext", "Werk", "SummevonSummevonZugang_Bedarf",
    "VERPR", "VERPR_vm", "cost_value", "cost_pct", "cost_part",
    "Bestellmengeneinheit", "ZPLP1", "ZPLP2", "ZPLP3", "GEWEI", "GROES",
    "NTGEW", "MSTAE", "WRKST", "ZZ_SPEC_ID", "VPRSV",
]


def cost_report(ups: list[dict], downs: list[dict]) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "cost_ups_top_500"
    _write_sheet(ws1, COST_HEADERS, ups,
                 currency_cols=("VERPR", "VERPR_vm", "cost_value", "cost_part"),
                 pct_cols=("cost_pct",))
    ws2 = wb.create_sheet("cost_downs_top_500")
    _write_sheet(ws2, COST_HEADERS, downs,
                 currency_cols=("VERPR", "VERPR_vm", "cost_value", "cost_part"),
                 pct_cols=("cost_pct",))
    return _to_bytes(wb)


def generic_report(sheet_name: str, rows: list[dict],
                   currency_cols=(), pct_cols=()) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"
    headers = list(rows[0].keys()) if rows else ["(no rows)"]
    _write_sheet(ws, headers, rows, currency_cols=currency_cols, pct_cols=pct_cols)
    return _to_bytes(wb)


def two_sheet_report(name_a: str, rows_a: list[dict],
                     name_b: str, rows_b: list[dict],
                     currency_cols=()) -> bytes:
    wb = Workbook()
    ws_a = wb.active
    ws_a.title = name_a[:31]
    _write_sheet(ws_a, list(rows_a[0].keys()) if rows_a else ["(no rows)"],
                 rows_a, currency_cols=currency_cols)
    ws_b = wb.create_sheet(name_b[:31])
    _write_sheet(ws_b, list(rows_b[0].keys()) if rows_b else ["(no rows)"],
                 rows_b, currency_cols=currency_cols)
    return _to_bytes(wb)
